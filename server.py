from fastapi import FastAPI, APIRouter, HTTPException
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pymongo import ReturnDocument
import os
import io
import json
import base64
import logging
import uuid
import re
import google.generativeai as genai
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.environ.get("MONGO_URL", "")
DB_NAME = os.environ.get("DB_NAME", "svl_docs")
GEMINI_API_KEY = os.environ.get("GEMINI_API_KEY", "")

if not MONGO_URL:
    raise RuntimeError("MONGO_URL is not configured in backend/.env")

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI(title="SVL Docs Scanner API", version="2.2.0")
api = APIRouter(prefix="/api")


# =========================
# Models
# =========================
class OCRRequest(BaseModel):
    image_base64: str
    mime_type: str = "image/jpeg"

    # V2.0 options used by frontend
    pdf_mode: bool = False
    detect_only_eway: bool = True
    auto_crop: bool = True
    ai_enhance: bool = True


class PDFOCRRequest(BaseModel):
    pdf_base64: str
    file_name: str = "document.pdf"
    detect_only_eway: bool = True


class BatchOCRItem(BaseModel):
    image_base64: str
    mime_type: str = "image/jpeg"
    file_name: str = ""


class BatchOCRRequest(BaseModel):
    items: List[BatchOCRItem] = Field(default_factory=list)
    detect_only_eway: bool = True
    auto_crop: bool = True
    ai_enhance: bool = True


class OCRField(BaseModel):
    value: str = ""
    confidence: float = 0.0


class OCRResult(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    document_type: str = "unknown"
    raw_text: str = ""
    fields: Dict[str, OCRField] = Field(default_factory=dict)
    overall_confidence: float = 0.0
    page_count: int = 0
    detected_pages: int = 0
    pages: List[Dict[str, Any]] = Field(default_factory=list)


class RecordCreate(BaseModel):
    # === CORE FIELDS ===
    date: str = ""
    lr_number: str = ""
    vehicle_number: str = ""
    supplier_gstin: str = ""
    supplier_name: str = ""
    dispatch_place: str = ""
    recipient_gstin: str = ""
    recipient_name: str = ""
    delivery_place: str = ""
    amount: str = ""

    # === V2.0 E-Way / PDF fields ===
    eway_bill_no: str = ""
    transporter: str = ""
    source_type: str = "image"
    mime_type: str = "image/jpeg"
    file_name: str = ""
    page_count: int = 0
    detected_pages: int = 0
    selected_page: int = 0
    pages: List[Dict[str, Any]] = Field(default_factory=list)

    # Recently Deleted
    is_deleted: bool = False
    deleted_at: str = ""

    # === Legacy / retained ===
    document_type: str = "lr_receipt"
    document_number: str = ""
    gstin: str = ""
    transporter_name: str = ""
    consignor: str = ""
    consignee: str = ""
    weight: str = ""
    distance: str = ""
    origin_address: str = ""
    destination_address: str = ""
    origin_city: str = ""
    destination_city: str = ""
    origin_state: str = ""
    destination_state: str = ""
    origin_pin: str = ""
    destination_pin: str = ""
    source: str = ""
    destination: str = ""
    remarks: str = ""
    verification_status: str = "pending"
    raw_text: str = ""
    image_base64: str = ""
    original_image_base64: str = ""
    processed_image_base64: str = ""
    pdf_base64: str = ""
    image_hash: str = ""
    file_paths: Dict[str, str] = Field(default_factory=dict)
    folder_id: Optional[str] = None
    confidence_scores: Dict[str, float] = Field(default_factory=dict)
    is_favorite: bool = False


class Record(RecordCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    serial_no: int = 0
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    updated_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    year: int = 0
    month: int = 0
    week: int = 0
    ocr_version: str = "SVL AI Engine 2.2"
    ocr_confidence: float = 0.0
    processed_at: str = ""
    status: str = "verified"
    review_required: bool = False


class RecordUpdate(BaseModel):
    date: Optional[str] = None
    lr_number: Optional[str] = None
    vehicle_number: Optional[str] = None
    supplier_gstin: Optional[str] = None
    supplier_name: Optional[str] = None
    dispatch_place: Optional[str] = None
    recipient_gstin: Optional[str] = None
    recipient_name: Optional[str] = None
    delivery_place: Optional[str] = None
    amount: Optional[str] = None

    eway_bill_no: Optional[str] = None
    transporter: Optional[str] = None
    source_type: Optional[str] = None
    mime_type: Optional[str] = None
    file_name: Optional[str] = None
    page_count: Optional[int] = None
    detected_pages: Optional[int] = None
    selected_page: Optional[int] = None
    pages: Optional[List[Dict[str, Any]]] = None
    is_deleted: Optional[bool] = None
    deleted_at: Optional[str] = None

    document_type: Optional[str] = None
    document_number: Optional[str] = None
    gstin: Optional[str] = None
    transporter_name: Optional[str] = None
    consignor: Optional[str] = None
    consignee: Optional[str] = None
    weight: Optional[str] = None
    distance: Optional[str] = None
    origin_address: Optional[str] = None
    destination_address: Optional[str] = None
    origin_city: Optional[str] = None
    destination_city: Optional[str] = None
    origin_state: Optional[str] = None
    destination_state: Optional[str] = None
    origin_pin: Optional[str] = None
    destination_pin: Optional[str] = None
    source: Optional[str] = None
    destination: Optional[str] = None
    remarks: Optional[str] = None
    verification_status: Optional[str] = None
    folder_id: Optional[str] = None
    is_favorite: Optional[bool] = None
    raw_text: Optional[str] = None
    image_base64: Optional[str] = None
    original_image_base64: Optional[str] = None
    processed_image_base64: Optional[str] = None
    pdf_base64: Optional[str] = None
    image_hash: Optional[str] = None
    file_paths: Optional[Dict[str, str]] = None
    confidence_scores: Optional[Dict[str, float]] = None


class FolderCreate(BaseModel):
    name: str
    color: str = "#FFB300"


class Folder(FolderCreate):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: str = Field(default_factory=lambda: datetime.now(timezone.utc).isoformat())
    record_count: int = 0


class ExportRequest(BaseModel):
    record_ids: Optional[List[str]] = None
    format: str = "csv"
    folder_id: Optional[str] = None
    scope: Optional[str] = None
    year: Optional[int] = None
    month: Optional[int] = None
    week: Optional[int] = None
    document_type: Optional[str] = None


class DuplicateRequest(BaseModel):
    lr_number: Optional[str] = ""
    document_number: Optional[str] = ""
    eway_bill_no: Optional[str] = ""
    vehicle_number: Optional[str] = ""
    supplier_gstin: Optional[str] = ""
    recipient_gstin: Optional[str] = ""
    gstin: Optional[str] = ""
    amount: Optional[str] = ""
    date: Optional[str] = ""
    image_hash: Optional[str] = ""
    exclude_id: Optional[str] = None


class ManualRecordRequest(BaseModel):
    date: str = ""
    lr_number: str = ""
    vehicle_number: str = ""
    supplier_gstin: str = ""
    supplier_name: str = ""
    dispatch_place: str = ""
    recipient_gstin: str = ""
    recipient_name: str = ""
    delivery_place: str = ""
    amount: str = ""
    transporter: str = ""
    remarks: str = ""
    folder_id: Optional[str] = None


# =========================
# OCR
# =========================
SYSTEM_PROMPT = """You are an expert OCR + extraction engine for Indian transport & logistics documents, especially E-Way Bills.

Return ONLY valid JSON. No prose, no markdown fences.
Schema:
{
 "document_type": "eway_bill|lr_receipt|bilty|gst_invoice|delivery_challan|invoice|unknown",
 "raw_text": "full raw OCR text",
 "fields": {
   "eway_bill_no": {"value": "", "confidence": 0.0},
   "date": {"value": "", "confidence": 0.0},
   "lr_number": {"value": "", "confidence": 0.0},
   "vehicle_number": {"value": "", "confidence": 0.0},
   "supplier_gstin": {"value": "", "confidence": 0.0},
   "supplier_name": {"value": "", "confidence": 0.0},
   "dispatch_place": {"value": "", "confidence": 0.0},
   "recipient_gstin": {"value": "", "confidence": 0.0},
   "recipient_name": {"value": "", "confidence": 0.0},
   "delivery_place": {"value": "", "confidence": 0.0},
   "amount": {"value": "", "confidence": 0.0},
   "transporter": {"value": "", "confidence": 0.0}
 },
 "overall_confidence": 0.0
}

Strict rules:
- If document is not readable, return document_type:"unknown".
- Never guess GSTIN. Never guess vehicle number. Never invent values.
- Leave field empty if not clearly found.
- GSTIN is 15-character alphanumeric.
- amount/value should be plain number string without currency symbols or commas.
- For E-Way Bill, eway_bill_no may be printed as E-Way Bill No, EWB No, EWB.
- date should be printed date, usually dd/mm/yyyy or dd-mm-yyyy.
"""


PDF_PROMPT = """You are an OCR engine for multi-page Indian transport PDFs.

Read ALL pages.
Detect only pages that are clearly E-Way Bills.
Ignore unrelated invoices, bilty pages, transport receipts, and blank pages when detect_only_eway is true.

Return ONLY valid JSON:
{
 "document_type": "eway_bill",
 "raw_text": "combined raw OCR text",
 "page_count": 0,
 "detected_pages": 0,
 "pages": [
   {
     "page": 1,
     "is_eway_bill": true,
     "confidence": 0.0,
     "fields": {
       "eway_bill_no": {"value": "", "confidence": 0.0},
       "date": {"value": "", "confidence": 0.0},
       "lr_number": {"value": "", "confidence": 0.0},
       "vehicle_number": {"value": "", "confidence": 0.0},
       "supplier_gstin": {"value": "", "confidence": 0.0},
       "supplier_name": {"value": "", "confidence": 0.0},
       "dispatch_place": {"value": "", "confidence": 0.0},
       "recipient_gstin": {"value": "", "confidence": 0.0},
       "recipient_name": {"value": "", "confidence": 0.0},
       "delivery_place": {"value": "", "confidence": 0.0},
       "amount": {"value": "", "confidence": 0.0},
       "transporter": {"value": "", "confidence": 0.0}
     }
   }
 ],
 "fields": {},
 "overall_confidence": 0.0
}

Do not invent missing values.
"""


DEFAULT_FIELD_KEYS = [
    "eway_bill_no", "date", "lr_number", "vehicle_number",
    "supplier_gstin", "supplier_name", "dispatch_place",
    "recipient_gstin", "recipient_name", "delivery_place",
    "amount", "transporter",
]


def _extract_json(s: str) -> Dict[str, Any]:
    s = (s or "").strip()
    if s.startswith("```"):
        s = re.sub(r"^```(?:json)?", "", s).strip()
        if s.endswith("```"):
            s = s[:-3].strip()
    try:
        return json.loads(s)
    except Exception:
        m = re.search(r"\{.*\}", s, re.DOTALL)
        if m:
            return json.loads(m.group(0))
        raise


def _normalize_fields(fields: Any) -> Dict[str, Dict[str, Any]]:
    if not isinstance(fields, dict):
        fields = {}

    out: Dict[str, Dict[str, Any]] = {}
    for key in DEFAULT_FIELD_KEYS:
        v = fields.get(key, {})
        if not isinstance(v, dict):
            v = {"value": str(v or ""), "confidence": 0.0}

        try:
            conf = float(v.get("confidence", 0.0) or 0.0)
        except Exception:
            conf = 0.0

        out[key] = {
            "value": str(v.get("value", "") or ""),
            "confidence": max(0.0, min(1.0, conf)),
        }

    # fallbacks
    if not out["eway_bill_no"]["value"] and fields.get("document_number"):
        dn = fields.get("document_number")
        if isinstance(dn, dict):
            out["eway_bill_no"] = {"value": str(dn.get("value", "") or ""), "confidence": float(dn.get("confidence", 0.0) or 0.0)}
    return out


def calculate_confidence(fields: Dict[str, Any]) -> float:
    if not fields:
        return 0.0

    scores: List[float] = []

    for value in fields.values():
        try:
            if isinstance(value, dict):
                if value.get("value"):
                    scores.append(float(value.get("confidence", 0.0) or 0.0))
            elif isinstance(value, (int, float)):
                scores.append(float(value))
        except Exception:
            continue

    if not scores:
        return 0.0

    normalized = [
        max(0.0, min(1.0, score / 100 if score > 1 else score))
        for score in scores
    ]
    return round(sum(normalized) / len(normalized), 4)



async def run_ocr(image_b64: str, mime_type: str) -> Dict[str, Any]:
    if not GEMINI_API_KEY:
        raise HTTPException(500, "Gemini API key not configured")

    genai.configure(api_key=GEMINI_API_KEY)
    model = genai.GenerativeModel("gemini-2.5-flash")

    try:
        file_bytes = base64.b64decode(image_b64)
        prompt = PDF_PROMPT if mime_type == "application/pdf" else SYSTEM_PROMPT

        response = model.generate_content(
            [
                prompt,
                {"mime_type": mime_type, "data": file_bytes},
            ],
            generation_config={"temperature": 0},
        )

        data = _extract_json(response.text or "{}")

        if mime_type == "application/pdf":
            pages = data.get("pages", []) if isinstance(data.get("pages", []), list) else []
            normalized_pages = []
            for idx, p in enumerate(pages, start=1):
                p = p if isinstance(p, dict) else {}
                normalized_pages.append({
                    "page": int(p.get("page") or idx),
                    "is_eway_bill": bool(p.get("is_eway_bill", True)),
                    "confidence": float(p.get("confidence", 0.0) or 0.0),
                    "fields": _normalize_fields(p.get("fields", {})),
                })
            first_fields = normalized_pages[0]["fields"] if normalized_pages else _normalize_fields({})
            return {
                "document_type": "eway_bill" if normalized_pages else "unknown",
                "raw_text": str(data.get("raw_text", "") or ""),
                "fields": first_fields,
                "overall_confidence": calculate_confidence(first_fields),
                "page_count": int(data.get("page_count") or len(normalized_pages)),
                "detected_pages": int(data.get("detected_pages") or len(normalized_pages)),
                "pages": normalized_pages,
            }

        fields = _normalize_fields(data.get("fields", {}))
        return {
            "document_type": str(data.get("document_type", "unknown") or "unknown"),
            "raw_text": str(data.get("raw_text", "") or ""),
            "fields": fields,
            "overall_confidence": float(data.get("overall_confidence", calculate_confidence(fields)) or 0.0),
        }

    except HTTPException:
        raise
    except Exception as e:
        logging.exception("Gemini OCR failed")
        raise HTTPException(500, f"OCR failed: {e}")


def _iso_now() -> datetime:
    return datetime.now(timezone.utc)


def _now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def _model_dump(model: BaseModel) -> Dict[str, Any]:
    return model.model_dump() if hasattr(model, "model_dump") else model.dict()


def _scope_query(scope: Optional[str], year: Optional[int], month: Optional[int], week: Optional[int]) -> Dict[str, Any]:
    now = _iso_now()
    if scope == "today":
        start = now.replace(hour=0, minute=0, second=0, microsecond=0).isoformat()
        return {"created_at": {"$gte": start}}
    if scope == "week":
        from datetime import timedelta
        start = (now - timedelta(days=7)).isoformat()
        return {"created_at": {"$gte": start}}
    if scope == "month":
        start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        return {"created_at": {"$gte": start}}
    if scope == "year":
        start = now.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0).isoformat()
        return {"created_at": {"$gte": start}}
    if year and week:
        from datetime import date, timedelta
        try:
            monday = date.fromisocalendar(year, week, 1)
            next_monday = monday + timedelta(days=7)
            return {"created_at": {"$gte": monday.isoformat(), "$lt": next_monday.isoformat()}}
        except Exception:
            return {}
    if year and month:
        y_start = f"{year:04d}-{month:02d}-01T00:00:00"
        nm = month + 1
        ny = year
        if nm > 12:
            nm = 1
            ny = year + 1
        y_end = f"{ny:04d}-{nm:02d}-01T00:00:00"
        return {"created_at": {"$gte": y_start, "$lt": y_end}}
    if year:
        return {"created_at": {"$gte": f"{year:04d}-01-01T00:00:00", "$lt": f"{year + 1:04d}-01-01T00:00:00"}}
    return {}


def _clean_upper(value: str) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper().strip()


def _clean_amount(value: str) -> str:
    cleaned = re.sub(r"[^0-9.]", "", str(value or ""))
    if cleaned.count(".") > 1:
        first, *rest = cleaned.split(".")
        cleaned = first + "." + "".join(rest)
    return cleaned


async def _next_serial_no() -> int:
    counter = await db.counters.find_one_and_update(
        {"_id": "records_serial"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )

    seq = int((counter or {}).get("seq", 0) or 0)
    if seq > 0:
        return seq

    last = await db.records.find_one(
        {},
        {"_id": 0, "serial_no": 1},
        sort=[("serial_no", -1)],
    )
    base = int((last or {}).get("serial_no") or 0)

    counter = await db.counters.find_one_and_update(
        {"_id": "records_serial"},
        {"$set": {"seq": base + 1}},
        upsert=True,
        return_document=ReturnDocument.AFTER,
    )
    return int(counter["seq"])


# =========================
# Routes
# =========================
@api.get("/")
async def root():
    return {"message": "SVL Docs Scanner API", "version": "2.2.0"}


@api.get("/health")
async def health():
    try:
        await db.command("ping")
        return {
            "ok": True,
            "version": "2.2.0",
            "database": "connected",
            "gemini_configured": bool(GEMINI_API_KEY),
        }
    except Exception as exc:
        raise HTTPException(
            status_code=503,
            detail=f"Database unavailable: {exc}",
        )


@api.get("/stats")
async def stats():
    active_q = {"is_deleted": {"$ne": True}}
    total = await db.records.count_documents(active_q)
    favorites = await db.records.count_documents({**active_q, "is_favorite": True})
    folders = await db.folders.count_documents({})
    pdf_records = await db.records.count_documents({**active_q, "source_type": "pdf"})
    image_records = await db.records.count_documents(
        {**active_q, "source_type": {"$nin": ["pdf", "manual"]}}
    )
    manual_records = await db.records.count_documents(
        {**active_q, "source_type": "manual"}
    )
    trash_records = await db.records.count_documents({"is_deleted": True})

    recent = await db.records.find(
        active_q,
        {"_id": 0, "image_base64": 0, "original_image_base64": 0, "processed_image_base64": 0, "pdf_base64": 0},
    ).sort("created_at", -1).limit(5).to_list(5)

    pipeline = [
        {"$match": active_q},
        {"$group": {"_id": "$document_type", "count": {"$sum": 1}}},
    ]
    by_type_cursor = db.records.aggregate(pipeline)
    by_type = {}
    async for row in by_type_cursor:
        by_type[row["_id"] or "unknown"] = row["count"]

    return {
        "total_records": total,
        "favorites": favorites,
        "folders": folders,
        "pdf_records": pdf_records,
        "image_records": image_records,
        "manual_records": manual_records,
        "trash_records": trash_records,
        "by_type": by_type,
        "recent": recent,
    }


@api.post("/ocr/scan")
async def ocr_scan(req: OCRRequest):
    if not req.image_base64:
        raise HTTPException(status_code=400, detail="image_base64 required")
    result = await run_ocr(req.image_base64, req.mime_type)
    return JSONResponse(status_code=200, content=result)


@api.post("/ocr/pdf")
async def scan_pdf(req: PDFOCRRequest):
    if not req.pdf_base64:
        raise HTTPException(status_code=400, detail="pdf_base64 required")
    result = await run_ocr(req.pdf_base64, "application/pdf")
    result["file_name"] = req.file_name
    return JSONResponse(status_code=200, content=result)


@api.post("/ocr/batch")
async def scan_batch(req: BatchOCRRequest):
    if not req.items:
        raise HTTPException(status_code=400, detail="items required")

    results: List[Dict[str, Any]] = []
    for index, item in enumerate(req.items, start=1):
        try:
            result = await run_ocr(item.image_base64, item.mime_type)
            result["index"] = index
            result["file_name"] = item.file_name
            results.append(result)
        except HTTPException as exc:
            results.append({
                "index": index,
                "file_name": item.file_name,
                "document_type": "unknown",
                "raw_text": "",
                "fields": _normalize_fields({}),
                "overall_confidence": 0.0,
                "error": str(exc.detail),
            })

    if req.detect_only_eway:
        results = [
            item for item in results
            if str(item.get("document_type", "")).lower() == "eway_bill"
        ]

    return {
        "items": results,
        "count": len(results),
        "processed": len(req.items),
    }


@api.post("/records", response_model=Record)
async def create_record(payload: RecordCreate):
    payload_data = _model_dump(payload)
    rec = Record(**payload_data)

    # Backward/forward field mirroring
    if rec.eway_bill_no and not rec.document_number:
        rec.document_number = rec.eway_bill_no
    if rec.lr_number and not rec.document_number:
        rec.document_number = rec.lr_number
    if rec.supplier_gstin and not rec.gstin:
        rec.gstin = rec.supplier_gstin
    if rec.supplier_name and not rec.consignor:
        rec.consignor = rec.supplier_name
    if rec.recipient_name and not rec.consignee:
        rec.consignee = rec.recipient_name
    if rec.dispatch_place and not rec.source:
        rec.source = rec.dispatch_place
    if rec.delivery_place and not rec.destination:
        rec.destination = rec.delivery_place

    if rec.source_type == "manual" or rec.mime_type == "text/manual":
        rec.source_type = "manual"
        rec.mime_type = "text/manual"
        rec.document_type = "manual_entry"
        rec.verification_status = "verified"
        rec.status = "verified"
        rec.review_required = False
        rec.ocr_confidence = 1.0
    elif rec.mime_type == "application/pdf" or rec.pdf_base64:
        rec.source_type = "pdf"
        if rec.document_type in ("", "lr_receipt", "pdf"):
            rec.document_type = "eway_bill" if rec.eway_bill_no or rec.detected_pages else "pdf"
    elif rec.eway_bill_no:
        rec.document_type = "eway_bill"

    now = _iso_now()
    rec.year = now.year
    rec.month = now.month
    rec.week = now.isocalendar().week
    rec.processed_at = now.isoformat()
    if rec.source_type == "manual":
        rec.ocr_confidence = 1.0
        rec.review_required = False
    else:
        rec.ocr_confidence = calculate_confidence(rec.confidence_scores)
        rec.review_required = (
            rec.ocr_confidence < 0.80
            if rec.ocr_confidence <= 1
            else rec.ocr_confidence < 80
        )

    rec.serial_no = await _next_serial_no()

    doc = _model_dump(rec)
    await db.records.insert_one(doc)
    doc.pop("_id", None)
    return rec


@api.post("/records/manual", response_model=Record)
async def create_manual_record(payload: ManualRecordRequest):
    data = _model_dump(payload)

    if not (
        str(data.get("lr_number") or "").strip()
        or str(data.get("vehicle_number") or "").strip()
        or str(data.get("supplier_name") or "").strip()
    ):
        raise HTTPException(
            status_code=400,
            detail="LR Number, Vehicle Number, or Supplier Name is required",
        )

    now = _iso_now()

    rec = Record(
        date=str(data.get("date") or "").strip(),
        lr_number=str(data.get("lr_number") or "").strip(),
        vehicle_number=_clean_upper(data.get("vehicle_number") or ""),
        supplier_gstin=_clean_upper(data.get("supplier_gstin") or "")[:15],
        supplier_name=str(data.get("supplier_name") or "").strip(),
        dispatch_place=str(data.get("dispatch_place") or "").strip(),
        recipient_gstin=_clean_upper(data.get("recipient_gstin") or "")[:15],
        recipient_name=str(data.get("recipient_name") or "").strip(),
        delivery_place=str(data.get("delivery_place") or "").strip(),
        amount=_clean_amount(data.get("amount") or ""),
        transporter=str(data.get("transporter") or "").strip(),
        transporter_name=str(data.get("transporter") or "").strip(),
        remarks=str(data.get("remarks") or "").strip(),
        folder_id=data.get("folder_id"),
        document_type="manual_entry",
        document_number=str(data.get("lr_number") or "").strip(),
        gstin=_clean_upper(data.get("supplier_gstin") or "")[:15],
        consignor=str(data.get("supplier_name") or "").strip(),
        consignee=str(data.get("recipient_name") or "").strip(),
        source=str(data.get("dispatch_place") or "").strip(),
        destination=str(data.get("delivery_place") or "").strip(),
        source_type="manual",
        mime_type="text/manual",
        verification_status="verified",
        status="verified",
        review_required=False,
        ocr_confidence=1.0,
        processed_at=now.isoformat(),
        year=now.year,
        month=now.month,
        week=now.isocalendar().week,
        ocr_version="SVL Manual Entry 2.2",
        confidence_scores={},
    )

    rec.serial_no = await _next_serial_no()

    doc = _model_dump(rec)
    await db.records.insert_one(doc)
    doc.pop("_id", None)
    return rec


@api.get("/records")
async def list_records(
    q: Optional[str] = None,
    folder_id: Optional[str] = None,
    favorites: Optional[bool] = None,
    document_type: Optional[str] = None,
    scope: Optional[str] = None,
    year: Optional[int] = None,
    month: Optional[int] = None,
    week: Optional[int] = None,
    limit: int = 200,
):
    if scope == "trash":
        query: Dict[str, Any] = {"is_deleted": True}
    else:
        query = {"is_deleted": {"$ne": True}}

    if folder_id:
        query["folder_id"] = folder_id
    if favorites:
        query["is_favorite"] = True
    if document_type:
        if document_type == "pdf":
            query["source_type"] = "pdf"
        elif document_type in ("manual", "manual_entry"):
            query["source_type"] = "manual"
        else:
            query["document_type"] = document_type

    if scope and scope != "trash":
        query.update(_scope_query(scope, year, month, week))
    else:
        query.update(_scope_query(None, year, month, week))

    if q:
        rx = {"$regex": re.escape(q), "$options": "i"}
        query["$or"] = [
            {"eway_bill_no": rx}, {"lr_number": rx}, {"vehicle_number": rx},
            {"supplier_gstin": rx}, {"supplier_name": rx}, {"dispatch_place": rx},
            {"recipient_gstin": rx}, {"recipient_name": rx}, {"delivery_place": rx},
            {"amount": rx}, {"date": rx},
            {"document_number": rx}, {"gstin": rx}, {"transporter": rx}, {"transporter_name": rx},
            {"consignor": rx}, {"consignee": rx}, {"source": rx}, {"destination": rx},
            {"remarks": rx}, {"raw_text": rx}, {"file_name": rx},
        ]

    cursor = db.records.find(
        query,
        {"_id": 0, "image_base64": 0, "original_image_base64": 0, "processed_image_base64": 0, "pdf_base64": 0},
    ).sort("created_at", -1).limit(limit)
    items = await cursor.to_list(limit)
    return {"items": items, "count": len(items)}


@api.get("/records/trash")
async def list_deleted_records(limit: int = 200):
    items = await db.records.find(
        {"is_deleted": True},
        {"_id": 0, "image_base64": 0, "original_image_base64": 0, "processed_image_base64": 0, "pdf_base64": 0},
    ).sort("deleted_at", -1).limit(limit).to_list(limit)
    return {"items": items, "count": len(items)}


@api.delete("/records/trash")
async def empty_trash():
    result = await db.records.delete_many({"is_deleted": True})
    return {"ok": True, "success": True, "deleted_count": result.deleted_count}


@api.get("/records/{record_id}")
async def get_record(record_id: str):
    rec = await db.records.find_one({"id": record_id}, {"_id": 0})
    if not rec:
        raise HTTPException(404, "Record not found")
    return rec


@api.patch("/records/{record_id}")
async def update_record(record_id: str, patch: RecordUpdate):
    updates = {k: v for k, v in _model_dump(patch).items() if v is not None}
    if not updates:
        rec = await db.records.find_one({"id": record_id}, {"_id": 0})
        if not rec:
            raise HTTPException(404, "Record not found")
        return rec

    if updates.get("eway_bill_no") and not updates.get("document_number"):
        updates["document_number"] = updates["eway_bill_no"]
    if updates.get("lr_number") and not updates.get("document_number"):
        updates["document_number"] = updates["lr_number"]
    if updates.get("supplier_gstin") and not updates.get("gstin"):
        updates["gstin"] = updates["supplier_gstin"]
    if updates.get("supplier_name") and not updates.get("consignor"):
        updates["consignor"] = updates["supplier_name"]
    if updates.get("recipient_name") and not updates.get("consignee"):
        updates["consignee"] = updates["recipient_name"]
    if updates.get("dispatch_place") and not updates.get("source"):
        updates["source"] = updates["dispatch_place"]
    if updates.get("delivery_place") and not updates.get("destination"):
        updates["destination"] = updates["delivery_place"]
    if updates.get("mime_type") == "application/pdf" or updates.get("pdf_base64"):
        updates["source_type"] = "pdf"

    updates["updated_at"] = _now_iso()
    res = await db.records.update_one({"id": record_id}, {"$set": updates})
    if res.matched_count == 0:
        raise HTTPException(404, "Record not found")
    rec = await db.records.find_one({"id": record_id}, {"_id": 0})
    return rec


@api.delete("/records/{record_id}")
async def delete_record(record_id: str):
    res = await db.records.update_one(
        {"id": record_id},
        {"$set": {"is_deleted": True, "deleted_at": _now_iso(), "updated_at": _now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Record not found")
    return {"ok": True, "success": True}


@api.post("/records/{record_id}/restore")
async def restore_record(record_id: str):
    res = await db.records.update_one(
        {"id": record_id},
        {"$set": {"is_deleted": False, "deleted_at": "", "updated_at": _now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Record not found")
    return {"ok": True, "success": True}


@api.delete("/records/{record_id}/forever")
async def delete_record_forever(record_id: str):
    res = await db.records.delete_one({"id": record_id})
    if res.deleted_count == 0:
        raise HTTPException(404, "Record not found")
    return {"ok": True, "success": True}


@api.delete("/records/{record_id}/permanent")
async def delete_record_permanent(record_id: str):
    return await delete_record_forever(record_id)


@api.post("/records/duplicates")
async def check_duplicates(payload: DuplicateRequest):
    data = _model_dump(payload)
    ors: List[Dict[str, Any]] = []

    ewb = (data.get("eway_bill_no") or "").strip()
    doc_no = (data.get("document_number") or data.get("lr_number") or "").strip()

    if ewb:
        ors.append({"eway_bill_no": ewb})
        ors.append({"document_number": ewb})
    if doc_no:
        ors.append({"lr_number": doc_no})
        ors.append({"document_number": doc_no})

    for key in ("vehicle_number", "supplier_gstin", "recipient_gstin", "gstin", "image_hash"):
        val = (data.get(key) or "").strip()
        if val:
            ors.append({key: val})

    amt = (data.get("amount") or "").strip()
    dt = (data.get("date") or "").strip()
    if amt and dt:
        ors.append({"$and": [{"amount": amt}, {"date": dt}]})

    if not ors:
        return {"duplicates": []}

    q: Dict[str, Any] = {"is_deleted": {"$ne": True}, "$or": ors}
    if data.get("exclude_id"):
        q["id"] = {"$ne": data["exclude_id"]}

    dups = await db.records.find(
        q,
        {"_id": 0, "image_base64": 0, "original_image_base64": 0, "processed_image_base64": 0, "pdf_base64": 0},
    ).limit(20).to_list(20)
    return {"duplicates": dups}


@api.post("/records/{record_id}/move")
async def move_record(record_id: str, payload: Dict[str, Any]):
    folder_id = payload.get("folder_id")
    res = await db.records.update_one(
        {"id": record_id},
        {"$set": {"folder_id": folder_id, "updated_at": _now_iso()}},
    )
    if res.matched_count == 0:
        raise HTTPException(404, "Record not found")
    return {"ok": True, "folder_id": folder_id}


# ----- Folders -----
@api.post("/folders", response_model=Folder)
async def create_folder(payload: FolderCreate):
    folder = Folder(**_model_dump(payload))
    await db.folders.insert_one(_model_dump(folder))
    return folder


@api.get("/folders")
async def list_folders():
    folders = await db.folders.find({}, {"_id": 0}).sort("created_at", -1).to_list(200)
    for f in folders:
        f["record_count"] = await db.records.count_documents({"folder_id": f["id"], "is_deleted": {"$ne": True}})
    return {"items": folders}


@api.get("/folders/smart")
async def smart_folders():
    active_q = {"is_deleted": {"$ne": True}}
    all_count = await db.records.count_documents(active_q)
    smart = []
    for key, label in [("today", "Today"), ("week", "This Week"), ("month", "This Month"), ("year", "This Year")]:
        q = {**active_q, **_scope_query(key, None, None, None)}
        smart.append({"key": key, "label": label, "count": await db.records.count_documents(q)})

    smart.extend([
        {"key": "all", "label": "All Records", "count": all_count},
        {"key": "pdf", "label": "PDF Records", "count": await db.records.count_documents({**active_q, "source_type": "pdf"})},
        {"key": "eway", "label": "E-Way Bills", "count": await db.records.count_documents({**active_q, "document_type": "eway_bill"})},
        {"key": "manual", "label": "Manual Entries", "count": await db.records.count_documents({**active_q, "source_type": "manual"})},
        {"key": "trash", "label": "Recently Deleted", "count": await db.records.count_documents({"is_deleted": True})},
    ])

    pipeline = [
        {"$match": active_q},
        {"$project": {
            "year": {"$substrBytes": ["$created_at", 0, 4]},
            "month": {"$substrBytes": ["$created_at", 5, 2]},
            "week": {"$ifNull": ["$week", 0]},
        }},
        {"$group": {
            "_id": {"year": "$year", "month": "$month", "week": "$week"},
            "count": {"$sum": 1},
        }},
        {"$sort": {"_id.year": -1, "_id.month": -1, "_id.week": -1}},
    ]
    year_map: Dict[str, Dict[str, Any]] = {}
    async for row in db.records.aggregate(pipeline):
        y = row["_id"]["year"]
        m = row["_id"]["month"]
        w = int(row["_id"].get("week") or 0)
        count = int(row["count"])

        year_entry = year_map.setdefault(
            y,
            {"year": y, "total": 0, "months": []},
        )

        month_entry = next(
            (item for item in year_entry["months"] if item["month"] == m),
            None,
        )
        if month_entry is None:
            month_entry = {"month": m, "count": 0, "weeks": []}
            year_entry["months"].append(month_entry)

        month_entry["count"] += count
        if w:
            month_entry["weeks"].append({"week": str(w), "count": count})
        year_entry["total"] += count

    return {"smart": smart, "years": list(year_map.values())}


@api.delete("/folders/{folder_id}")
async def delete_folder(folder_id: str):
    await db.folders.delete_one({"id": folder_id})
    await db.records.update_many({"folder_id": folder_id}, {"$set": {"folder_id": None, "updated_at": _now_iso()}})
    return {"ok": True}


# ----- Export -----
CORE_COLUMNS = [
    ("serial_no", "Sr No."),
    ("date", "Date"),
    ("lr_number", "LR Number"),
    ("vehicle_number", "Vehicle Number"),
    ("supplier_gstin", "Supplier GSTIN"),
    ("supplier_name", "Supplier Name"),
    ("dispatch_place", "Dispatch Place"),
    ("recipient_gstin", "Recipient GSTIN"),
    ("recipient_name", "Recipient Name"),
    ("delivery_place", "Delivery Place"),
    ("amount", "Amount"),
]


def _core_row(r: Dict[str, Any], idx: int) -> List[str]:
    row = []
    for key, _ in CORE_COLUMNS:
        if key == "serial_no":
            row.append(str(idx))
            continue

        v = r.get(key) or ""
        if not v:
            if key == "lr_number":
                v = r.get("lr_number") or r.get("document_number") or ""
            elif key == "supplier_gstin":
                v = r.get("supplier_gstin") or r.get("gstin") or ""
            elif key == "supplier_name":
                v = r.get("supplier_name") or r.get("consignor") or r.get("transporter_name") or ""
            elif key == "dispatch_place":
                v = r.get("dispatch_place") or r.get("source") or r.get("origin_city") or ""
            elif key == "recipient_name":
                v = r.get("recipient_name") or r.get("consignee") or ""
            elif key == "delivery_place":
                v = r.get("delivery_place") or r.get("destination") or r.get("destination_city") or ""
        row.append(str(v))
    return row


def _build_csv(records: List[Dict[str, Any]]) -> bytes:
    import csv
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([label for _, label in CORE_COLUMNS])
    for idx, r in enumerate(records, start=1):
        w.writerow(_core_row(r, idx))
    return buf.getvalue().encode("utf-8")


def _build_txt(records: List[Dict[str, Any]]) -> bytes:
    header = " | ".join(label for _, label in CORE_COLUMNS)
    lines = [header, "-" * len(header)]
    for idx, r in enumerate(records, start=1):
        lines.append(" | ".join(_core_row(r, idx)))
    return "\n".join(lines).encode("utf-8")


def _build_xlsx(records: List[Dict[str, Any]]) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "SVL Docs"

    labels = [label for _, label in CORE_COLUMNS]
    ws.append(labels)

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="0A57D8", end_color="0A57D8", fill_type="solid")
    thin = Side(border_style="thin", color="BBD7FF")
    border = Border(top=thin, left=thin, right=thin, bottom=thin)

    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for idx, r in enumerate(records, start=1):
        ws.append(_core_row(r, idx))
        for cell in ws[ws.max_row]:
            cell.border = border
            cell.alignment = Alignment(vertical="center", wrap_text=True)

    widths = [8, 14, 18, 18, 22, 26, 22, 22, 26, 22, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    ws.freeze_panes = "A2"
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def _build_pdf(records: List[Dict[str, Any]]) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors as rl_colors
    from reportlab.lib.units import mm

    bio = io.BytesIO()
    doc = SimpleDocTemplate(
        bio, pagesize=landscape(A4),
        leftMargin=8 * mm, rightMargin=8 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
    )

    styles = getSampleStyleSheet()
    story = [
        Paragraph("<b>SVL Docs Scanner — Records Ledger</b>", styles["Title"]),
        Paragraph(f"Generated: {datetime.now().strftime('%d %b %Y, %H:%M')}  |  Total: {len(records)}", styles["Normal"]),
        Spacer(1, 6),
    ]

    data = [[label for _, label in CORE_COLUMNS]]
    for idx, r in enumerate(records, start=1):
        data.append(_core_row(r, idx))

    col_widths = [12, 20, 26, 28, 34, 38, 32, 34, 38, 32, 22]
    tbl = Table(data, colWidths=[w * mm for w in col_widths], repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), rl_colors.HexColor("#0A57D8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), rl_colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [rl_colors.HexColor("#F4F8FF"), rl_colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.3, rl_colors.HexColor("#BBD7FF")),
        ("LEFTPADDING", (0, 0), (-1, -1), 3),
        ("RIGHTPADDING", (0, 0), (-1, -1), 3),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)

    if not records:
        story.append(Paragraph("<i>No records to export.</i>", styles["Normal"]))

    doc.build(story)
    return bio.getvalue()


@api.post("/export")
async def export_records(req: ExportRequest):
    query: Dict[str, Any] = {"is_deleted": {"$ne": True}}

    if req.record_ids:
        query["id"] = {"$in": req.record_ids}
    else:
        if req.folder_id:
            query["folder_id"] = req.folder_id
        if req.document_type:
            if req.document_type == "pdf":
                query["source_type"] = "pdf"
            elif req.document_type in ("manual", "manual_entry"):
                query["source_type"] = "manual"
            else:
                query["document_type"] = req.document_type
        query.update(_scope_query(req.scope, req.year, req.month, req.week))

    records = await db.records.find(
        query,
        {"_id": 0, "image_base64": 0, "original_image_base64": 0, "processed_image_base64": 0, "pdf_base64": 0},
    ).sort([
        ("lr_number", 1),
        ("document_number", 1),
        ("created_at", 1),
    ]).to_list(5000)

    fmt = (req.format or "csv").lower()
    if fmt == "csv":
        data, mime, ext = _build_csv(records), "text/csv", "csv"
    elif fmt == "json":
        data, mime, ext = json.dumps(records, indent=2, default=str).encode("utf-8"), "application/json", "json"
    elif fmt == "txt":
        data, mime, ext = _build_txt(records), "text/plain", "txt"
    elif fmt == "xlsx":
        data, mime, ext = _build_xlsx(records), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "xlsx"
    elif fmt == "pdf":
        data, mime, ext = _build_pdf(records), "application/pdf", "pdf"
    else:
        raise HTTPException(400, "invalid format")

    b64 = base64.b64encode(data).decode("utf-8")
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    return {"filename": f"svl-docs-{ts}.{ext}", "mime_type": mime, "base64": b64, "count": len(records)}


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()