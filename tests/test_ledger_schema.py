"""
DRCDocs Enterprise — Ledger schema (11-column) backend tests.

Focus (per review request):
 1. POST /api/records assigns auto-incrementing serial_no (>= previous max+1).
 2. POST /api/records with the new 10 fields persists correctly; lr_number
    is mirrored into legacy document_number.
 3. POST /api/records/duplicates matches by lr_number (and legacy
    document_number), vehicle_number, supplier_gstin, recipient_gstin,
    image_hash, amount+date. exclude_id suppresses self.
 4. PATCH /api/records/{id} — edit in place, serial_no unchanged, no new row.
 5. GET /api/records ?q= hits the new fields.
 6. POST /api/export xlsx — exact header order + row order by serial_no,
    first data row is the earliest inserted record.
 7. POST /api/export pdf — %PDF signature, empty + non-empty both render.
 8. POST /api/export csv & txt — header row + Sr No. column present.
 9. GET /api/folders/smart — still works and reflects counts.
10. POST /api/records/{id}/move — record_count on folder updates.

DOES NOT touch /api/ocr/scan.
"""
import os
import io
import base64
import pytest
import requests
from datetime import datetime, timezone

BASE_URL = os.environ.get(
    "EXPO_PUBLIC_BACKEND_URL",
    "https://logistics-ocr-pro-1.preview.emergentagent.com",
).rstrip("/")


EXPECTED_HEADERS = [
    "Sr No.", "Date", "LR Number", "Vehicle Number",
    "Supplier GSTIN", "Supplier Name", "Dispatch Place",
    "Recipient GSTIN", "Recipient Name", "Delivery Place", "Amount",
]


@pytest.fixture(scope="session")
def api():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    return s


@pytest.fixture(scope="session")
def state():
    return {"records": [], "folders": [], "serial_nos": []}


def _payload(idx: int) -> dict:
    return {
        "date": f"2026-01-{idx:02d}",
        "lr_number": f"TEST-LEDGER-LR-{idx}",
        "vehicle_number": f"MH14ZZ{idx:04d}",
        "supplier_gstin": f"27TESTSUP{idx:04d}Z5",
        "supplier_name": f"Test Supplier {idx}",
        "dispatch_place": f"Mumbai-{idx}",
        "recipient_gstin": f"29TESTREC{idx:04d}Z5",
        "recipient_name": f"Test Recipient {idx}",
        "delivery_place": f"Delhi-{idx}",
        "amount": f"{1000 + idx}.00",
        "image_hash": f"hash-ledger-{idx}",
    }


# ---------------------------------------------------------------------------
# 1) Serial number auto-increment across multiple inserts
# ---------------------------------------------------------------------------
class TestSerialAutoIncrement:
    def test_create_three_records_serial_monotonic(self, api, state):
        prev_max = 0
        for i in (1, 2, 3):
            r = api.post(f"{BASE_URL}/api/records", json=_payload(i))
            assert r.status_code == 200, r.text
            d = r.json()
            assert "serial_no" in d and isinstance(d["serial_no"], int)
            assert d["serial_no"] >= prev_max + 1, (
                f"serial_no did not advance: prev={prev_max} new={d['serial_no']}"
            )
            prev_max = d["serial_no"]
            state["records"].append(d["id"])
            state["serial_nos"].append(d["serial_no"])
        # Strictly increasing by exactly +1 for our 3 consecutive inserts
        assert state["serial_nos"][1] == state["serial_nos"][0] + 1
        assert state["serial_nos"][2] == state["serial_nos"][1] + 1

    def test_lr_number_mirrored_to_document_number(self, api, state):
        rid = state["records"][0]
        rec = api.get(f"{BASE_URL}/api/records/{rid}").json()
        assert rec["lr_number"] == "TEST-LEDGER-LR-1"
        assert rec["document_number"] == "TEST-LEDGER-LR-1", (
            f"document_number was not mirrored: {rec.get('document_number')!r}"
        )

    def test_new_fields_persisted(self, api, state):
        rid = state["records"][0]
        rec = api.get(f"{BASE_URL}/api/records/{rid}").json()
        p = _payload(1)
        for k in ("date", "lr_number", "vehicle_number",
                  "supplier_gstin", "supplier_name", "dispatch_place",
                  "recipient_gstin", "recipient_name", "delivery_place",
                  "amount"):
            assert rec[k] == p[k], f"{k} mismatch: {rec.get(k)!r} vs {p[k]!r}"
        assert "_id" not in rec


# ---------------------------------------------------------------------------
# 2) Duplicate detection over new fields (+ legacy fallback)
# ---------------------------------------------------------------------------
class TestDuplicatesNewSchema:
    def test_by_lr_number(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"lr_number": "TEST-LEDGER-LR-1"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_by_lr_number_legacy_document_number_fallback(self, api, state):
        # lr_number is mirrored into document_number, so a search with
        # lr_number field must also OR against document_number field.
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"lr_number": "TEST-LEDGER-LR-1"})
        assert r.status_code == 200
        dups = r.json()["duplicates"]
        # Since our record has BOTH lr_number and document_number set to the
        # same value, it should be matched (dedup via $or).
        assert any(d["id"] == rid for d in dups)

    def test_by_vehicle_number(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"vehicle_number": "MH14ZZ0001"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_by_supplier_gstin(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"supplier_gstin": "27TESTSUP0001Z5"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_by_recipient_gstin(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"recipient_gstin": "29TESTREC0001Z5"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_by_image_hash(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"image_hash": "hash-ledger-1"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_by_amount_and_date(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"amount": "1001.00", "date": "2026-01-01"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_amount_alone_returns_empty(self, api):
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"amount": "1001.00"})
        assert r.status_code == 200
        assert r.json()["duplicates"] == []

    def test_exclude_id_suppresses_self(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates", json={
            "lr_number": "TEST-LEDGER-LR-1",
            "vehicle_number": "MH14ZZ0001",
            "supplier_gstin": "27TESTSUP0001Z5",
            "recipient_gstin": "29TESTREC0001Z5",
            "image_hash": "hash-ledger-1",
            "exclude_id": rid,
        })
        assert r.status_code == 200
        assert all(d["id"] != rid for d in r.json()["duplicates"])


# ---------------------------------------------------------------------------
# 3) PATCH edit-in-place — serial_no unchanged
# ---------------------------------------------------------------------------
class TestPatchPreservesSerial:
    def test_patch_does_not_change_serial_no(self, api, state):
        rid = state["records"][0]
        before = api.get(f"{BASE_URL}/api/records/{rid}").json()
        before_sn = before["serial_no"]

        r = api.patch(f"{BASE_URL}/api/records/{rid}", json={
            "supplier_name": "Edited Supplier 1",
            "recipient_name": "Edited Recipient 1",
            "dispatch_place": "Edited Mumbai-1",
            "delivery_place": "Edited Delhi-1",
        })
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["supplier_name"] == "Edited Supplier 1"
        assert d["recipient_name"] == "Edited Recipient 1"
        assert d["dispatch_place"] == "Edited Mumbai-1"
        assert d["delivery_place"] == "Edited Delhi-1"
        assert d["serial_no"] == before_sn, (
            f"serial_no changed on PATCH: {before_sn} -> {d['serial_no']}"
        )

    def test_patch_does_not_create_duplicate_row(self, api, state):
        rid = state["records"][0]
        before = api.get(
            f"{BASE_URL}/api/records", params={"q": "TEST-LEDGER-LR-1"}
        ).json()["count"]
        r = api.patch(f"{BASE_URL}/api/records/{rid}",
                      json={"amount": "9999.00"})
        assert r.status_code == 200
        after = api.get(
            f"{BASE_URL}/api/records", params={"q": "TEST-LEDGER-LR-1"}
        ).json()["count"]
        assert before == after, "PATCH must not create duplicate rows"


# ---------------------------------------------------------------------------
# 4) Search across new fields
# ---------------------------------------------------------------------------
class TestSearchNewFields:
    def _found(self, api, q, rid):
        r = api.get(f"{BASE_URL}/api/records", params={"q": q})
        assert r.status_code == 200, r.text
        return rid in [x["id"] for x in r.json()["items"]]

    def test_search_by_lr_number(self, api, state):
        assert self._found(api, "TEST-LEDGER-LR-2", state["records"][1])

    def test_search_by_supplier_name(self, api, state):
        assert self._found(api, "Test Supplier 2", state["records"][1])

    def test_search_by_dispatch_place(self, api, state):
        assert self._found(api, "Mumbai-2", state["records"][1])

    def test_search_by_recipient_name(self, api, state):
        assert self._found(api, "Test Recipient 3", state["records"][2])

    def test_search_by_delivery_place(self, api, state):
        assert self._found(api, "Delhi-3", state["records"][2])

    def test_search_by_supplier_gstin(self, api, state):
        assert self._found(api, "27TESTSUP0002Z5", state["records"][1])

    def test_search_by_recipient_gstin(self, api, state):
        assert self._found(api, "29TESTREC0003Z5", state["records"][2])


# ---------------------------------------------------------------------------
# 5) Export XLSX — exact header order + row order by serial_no
# ---------------------------------------------------------------------------
class TestExportXlsx:
    def test_xlsx_headers_exact_order(self, api, state):
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "xlsx",
            "record_ids": state["records"],
        })
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["filename"].endswith(".xlsx")
        raw = base64.b64decode(d["base64"])
        assert raw[:2] == b"PK", "not a valid xlsx (zip) file"

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        header = [c.value for c in ws[1]]
        assert header == EXPECTED_HEADERS, (
            f"header order mismatch.\n"
            f"expected: {EXPECTED_HEADERS}\n"
            f"got:      {header}"
        )

    def test_xlsx_row_order_by_serial_no(self, api, state):
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "xlsx",
            "record_ids": state["records"],
        })
        assert r.status_code == 200
        raw = base64.b64decode(r.json()["base64"])
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        assert len(rows) >= 3
        serials = [int(row[0]) for row in rows]
        assert serials == sorted(serials), (
            f"rows not sorted by Sr No.: {serials}"
        )

    def test_xlsx_first_data_row_is_earliest_serial(self, api, state):
        # First inserted record has the smallest serial_no of our batch.
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "xlsx",
            "record_ids": state["records"],
        })
        assert r.status_code == 200
        raw = base64.b64decode(r.json()["base64"])
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        first_row = next(ws.iter_rows(min_row=2, max_row=2, values_only=True))
        # Fields expected for record[0] (post-patch)
        # Sr No. = state["serial_nos"][0]
        assert int(first_row[0]) == state["serial_nos"][0]
        assert first_row[1] == "2026-01-01"                       # Date
        assert first_row[2] == "TEST-LEDGER-LR-1"                # LR Number
        assert first_row[3] == "MH14ZZ0001"                      # Vehicle
        assert first_row[4] == "27TESTSUP0001Z5"                 # Sup GSTIN
        assert first_row[5] == "Edited Supplier 1"               # Sup Name
        assert first_row[6] == "Edited Mumbai-1"                 # Dispatch
        assert first_row[7] == "29TESTREC0001Z5"                 # Rec GSTIN
        assert first_row[8] == "Edited Recipient 1"              # Rec Name
        assert first_row[9] == "Edited Delhi-1"                  # Delivery
        assert first_row[10] == "9999.00"                        # Amount


# ---------------------------------------------------------------------------
# 6) Export CSV & TXT — header + Sr No.
# ---------------------------------------------------------------------------
class TestExportCsvTxt:
    def test_csv_header_and_sr_no(self, api, state):
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "csv",
            "record_ids": state["records"],
        })
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["filename"].endswith(".csv")
        raw = base64.b64decode(d["base64"]).decode("utf-8")
        lines = raw.strip().splitlines()
        # Header
        header = lines[0]
        assert header.split(",")[0] == "Sr No."
        for h in EXPECTED_HEADERS:
            assert h in header, f"missing header {h!r} in csv"
        # First data row: Sr No. column non-empty integer
        first_data = lines[1].split(",")
        assert first_data[0].strip().isdigit()

    def test_txt_header_and_sr_no(self, api, state):
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "txt",
            "record_ids": state["records"],
        })
        assert r.status_code == 200
        raw = base64.b64decode(r.json()["base64"]).decode("utf-8")
        lines = raw.splitlines()
        assert "Sr No." in lines[0]
        for h in EXPECTED_HEADERS:
            assert h in lines[0]


# ---------------------------------------------------------------------------
# 7) Export PDF — %PDF signature, empty + non-empty
# ---------------------------------------------------------------------------
class TestExportPdf:
    def test_pdf_non_empty_records(self, api, state):
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "pdf",
            "record_ids": state["records"],
        })
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["filename"].endswith(".pdf")
        assert d["mime_type"] == "application/pdf"
        raw = base64.b64decode(d["base64"])
        assert raw.startswith(b"%PDF"), "PDF signature missing"
        assert len(raw) > 500, "PDF unexpectedly small"

    def test_pdf_empty_record_set(self, api):
        # Force empty by passing a non-existent id list
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "pdf",
            "record_ids": ["definitely-does-not-exist-xyz"],
        })
        assert r.status_code == 200
        d = r.json()
        assert d["count"] == 0
        raw = base64.b64decode(d["base64"])
        assert raw.startswith(b"%PDF"), "empty PDF still needs %PDF sig"
        assert len(raw) > 200


# ---------------------------------------------------------------------------
# 8) Smart folders + move — record_count reflects moves
# ---------------------------------------------------------------------------
class TestFoldersSmartAndMove:
    def test_smart_shape_and_counts(self, api):
        r = api.get(f"{BASE_URL}/api/folders/smart")
        assert r.status_code == 200
        data = r.json()
        keys = {s["key"] for s in data["smart"]}
        assert {"today", "week", "month", "year", "all"}.issubset(keys)
        # all == /api/stats total_records
        all_row = next(s for s in data["smart"] if s["key"] == "all")
        stats = api.get(f"{BASE_URL}/api/stats").json()
        assert all_row["count"] == stats["total_records"]

    def test_create_folder_and_move_updates_count(self, api, state):
        # Create a folder
        r = api.post(f"{BASE_URL}/api/folders",
                     json={"name": "TEST_LEDGER_FOLDER", "color": "#0088FF"})
        assert r.status_code == 200, r.text
        fid = r.json()["id"]
        state["folders"].append(fid)

        # Baseline count
        before = api.get(f"{BASE_URL}/api/folders").json()["items"]
        before_target = next((f for f in before if f["id"] == fid), None)
        assert before_target is not None
        before_count = before_target["record_count"]

        # Move 2 of our records into it
        for rid in state["records"][:2]:
            mv = api.post(f"{BASE_URL}/api/records/{rid}/move",
                          json={"folder_id": fid})
            assert mv.status_code == 200

        # New count reflects moves
        after = api.get(f"{BASE_URL}/api/folders").json()["items"]
        after_target = next(f for f in after if f["id"] == fid)
        assert after_target["record_count"] == before_count + 2

    def test_unassign_folder_with_null(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/{rid}/move",
                     json={"folder_id": None})
        assert r.status_code == 200
        rec = api.get(f"{BASE_URL}/api/records/{rid}").json()
        assert rec["folder_id"] in (None, "")


# ---------------------------------------------------------------------------
# ZZ. Cleanup
# ---------------------------------------------------------------------------
class TestZZCleanup:
    def test_delete_records(self, api, state):
        for rid in list(state["records"]):
            r = api.delete(f"{BASE_URL}/api/records/{rid}")
            assert r.status_code == 200
        state["records"].clear()

    def test_delete_folders(self, api, state):
        for fid in list(state["folders"]):
            r = api.delete(f"{BASE_URL}/api/folders/{fid}")
            assert r.status_code == 200
        state["folders"].clear()
