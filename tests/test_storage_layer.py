"""
DRCDocs Enterprise — Storage / Edit / Folder / Duplicate / Export layer tests.

Focus (per review request):
- POST /api/records with new fields (source, destination, remarks, verification_status, image_hash)
- PATCH /api/records/{id} in-place edit (no duplicate, updated_at moves forward)
- POST /api/records/duplicates (doc no, vehicle no, gstin, image_hash, amount+date) with exclude_id
- POST /api/records/{id}/move (assign + unassign folder)
- GET /api/records `q` search across new fields (source/destination/remarks)
- GET /api/records scope filters (today/week/month/year, year+month)
- GET /api/folders/smart returns counts + year->months breakdown
- POST /api/export xlsx with scope filters — decode base64 and inspect workbook
- Folders create + list with record_count reflecting moves
- Persistence across restart (implicit: fetch records after inserts across separate requests)

Does NOT touch /api/ocr/scan (LLM cost).
"""
import os
import io
import time
import base64
import pytest
import requests
from datetime import datetime, timezone

BASE_URL = os.environ.get(
    "EXPO_PUBLIC_BACKEND_URL",
    "https://logistics-ocr-pro-1.preview.emergentagent.com",
).rstrip("/")


@pytest.fixture(scope="session")
def api():
    s = requests.Session()
    s.headers.update({"Content-Type": "application/json"})
    return s


@pytest.fixture(scope="session")
def state():
    return {"records": [], "folders": []}


# ---------------------------------------------------------------------------
# 1. POST /api/records — new fields persist and are returned
# ---------------------------------------------------------------------------
class TestCreateWithNewFields:
    def test_create_record_new_fields(self, api, state):
        payload = {
            "document_type": "invoice",
            "document_number": "TEST-STORE-1",
            "vehicle_number": "MH14ZZ0001",
            "gstin": "27TESTSTORE01",
            "transporter_name": "Store Transporter",
            "amount": "1500.00",
            "date": "2026-01-05",
            "source": "Mumbai",
            "destination": "Delhi",
            "remarks": "handle with care STORE-REMARK-KEYWORD",
            "verification_status": "verified",
            "image_hash": "hash-STORE-1",
        }
        r = api.post(f"{BASE_URL}/api/records", json=payload)
        assert r.status_code == 200, r.text
        d = r.json()
        # New fields all echoed
        for k in ("source", "destination", "remarks",
                  "verification_status", "image_hash"):
            assert d[k] == payload[k], f"{k} mismatch: {d.get(k)!r} vs {payload[k]!r}"
        assert d["verification_status"] == "verified"
        assert "id" in d and "_id" not in d
        assert "created_at" in d and "updated_at" in d
        state["records"].append(d["id"])

    def test_created_record_readable_via_get(self, api, state):
        rid = state["records"][0]
        r = api.get(f"{BASE_URL}/api/records/{rid}")
        assert r.status_code == 200
        d = r.json()
        assert d["source"] == "Mumbai"
        assert d["destination"] == "Delhi"
        assert d["remarks"].startswith("handle with care")
        assert d["image_hash"] == "hash-STORE-1"
        assert "_id" not in d

    def test_defaults_when_omitted(self, api, state):
        r = api.post(f"{BASE_URL}/api/records", json={
            "document_number": "TEST-STORE-2",
            "amount": "200.00",
            "date": "2026-01-06",
        })
        assert r.status_code == 200, r.text
        d = r.json()
        # Defaults: verification_status pending, other new fields empty strings
        assert d["verification_status"] == "pending"
        assert d["source"] == ""
        assert d["destination"] == ""
        assert d["remarks"] == ""
        assert d["image_hash"] == ""
        state["records"].append(d["id"])


# ---------------------------------------------------------------------------
# 2. PATCH /api/records/{id} — edit in place, updated_at advances, no duplicate
# ---------------------------------------------------------------------------
class TestPatchEditInPlace:
    def test_patch_updates_fields_and_updated_at(self, api, state):
        rid = state["records"][0]
        before = api.get(f"{BASE_URL}/api/records/{rid}").json()
        before_updated = before["updated_at"]
        before_created = before["created_at"]
        # Ensure clock moves
        time.sleep(1.1)
        r = api.patch(f"{BASE_URL}/api/records/{rid}", json={
            "remarks": "edited-remark",
            "verification_status": "unverified",
            "source": "Pune",
        })
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["remarks"] == "edited-remark"
        assert d["verification_status"] == "unverified"
        assert d["source"] == "Pune"
        # id unchanged, created_at unchanged, updated_at moved forward
        assert d["id"] == rid
        assert d["created_at"] == before_created
        assert d["updated_at"] > before_updated, (
            f"updated_at did not advance: {before_updated} -> {d['updated_at']}"
        )

    def test_patch_does_not_create_duplicate(self, api, state):
        rid = state["records"][0]
        # Count records with this document_number before and after another patch
        before = api.get(
            f"{BASE_URL}/api/records", params={"q": "TEST-STORE-1"}
        ).json()["count"]
        r = api.patch(f"{BASE_URL}/api/records/{rid}", json={"amount": "1600.00"})
        assert r.status_code == 200
        after = api.get(
            f"{BASE_URL}/api/records", params={"q": "TEST-STORE-1"}
        ).json()["count"]
        assert before == after, "PATCH should not create a duplicate record"

    def test_patch_404_for_unknown(self, api):
        r = api.patch(f"{BASE_URL}/api/records/does-not-exist",
                      json={"remarks": "x"})
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# 3. POST /api/records/duplicates — all keys + exclude_id
# ---------------------------------------------------------------------------
class TestDuplicates:
    def test_by_document_number(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"document_number": "TEST-STORE-1"})
        assert r.status_code == 200
        dups = r.json()["duplicates"]
        assert any(d["id"] == rid for d in dups)

    def test_by_vehicle_number(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"vehicle_number": "MH14ZZ0001"})
        assert r.status_code == 200
        dups = r.json()["duplicates"]
        assert any(d["id"] == rid for d in dups)

    def test_by_gstin(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"gstin": "27TESTSTORE01"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_by_image_hash(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"image_hash": "hash-STORE-1"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"])

    def test_by_amount_and_date(self, api, state):
        # record[0] has amount 1600.00 (after patch) date 2026-01-05
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"amount": "1600.00", "date": "2026-01-05"})
        assert r.status_code == 200
        assert any(d["id"] == rid for d in r.json()["duplicates"]), (
            f"amount+date combo did not find record; dups={r.json()['duplicates']}"
        )

    def test_amount_alone_does_not_match(self, api):
        # amount without date should NOT trigger amount-only match per spec
        r = api.post(f"{BASE_URL}/api/records/duplicates",
                     json={"amount": "1600.00"})
        assert r.status_code == 200
        # It should return empty (no OR clauses since amount+date requires both)
        assert r.json()["duplicates"] == []

    def test_exclude_id_suppresses_self(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/duplicates", json={
            "document_number": "TEST-STORE-1",
            "vehicle_number": "MH14ZZ0001",
            "gstin": "27TESTSTORE01",
            "image_hash": "hash-STORE-1",
            "exclude_id": rid,
        })
        assert r.status_code == 200
        dups = r.json()["duplicates"]
        assert all(d["id"] != rid for d in dups), (
            "exclude_id must suppress the record being edited"
        )


# ---------------------------------------------------------------------------
# 4. POST /api/folders + GET, and /api/records/{id}/move
# ---------------------------------------------------------------------------
class TestFoldersAndMove:
    def test_create_folder(self, api, state):
        r = api.post(f"{BASE_URL}/api/folders",
                     json={"name": "TEST_STORE_FOLDER", "color": "#00AAFF"})
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["name"] == "TEST_STORE_FOLDER"
        assert "id" in d and "_id" not in d
        state["folders"].append(d["id"])

    def test_move_record_to_folder(self, api, state):
        rid = state["records"][0]
        fid = state["folders"][0]
        r = api.post(f"{BASE_URL}/api/records/{rid}/move",
                     json={"folder_id": fid})
        assert r.status_code == 200
        body = r.json()
        assert body["ok"] is True
        assert body["folder_id"] == fid
        # verify via GET
        rec = api.get(f"{BASE_URL}/api/records/{rid}").json()
        assert rec["folder_id"] == fid

    def test_folder_record_count_reflects_move(self, api, state):
        fid = state["folders"][0]
        r = api.get(f"{BASE_URL}/api/folders")
        assert r.status_code == 200
        folders = r.json()["items"]
        target = next((f for f in folders if f["id"] == fid), None)
        assert target is not None, "folder not returned in list"
        assert target["record_count"] >= 1

    def test_list_records_by_folder(self, api, state):
        fid = state["folders"][0]
        rid = state["records"][0]
        r = api.get(f"{BASE_URL}/api/records", params={"folder_id": fid})
        assert r.status_code == 200
        ids = [x["id"] for x in r.json()["items"]]
        assert rid in ids

    def test_unassign_folder_with_null(self, api, state):
        rid = state["records"][0]
        r = api.post(f"{BASE_URL}/api/records/{rid}/move",
                     json={"folder_id": None})
        assert r.status_code == 200
        assert r.json()["folder_id"] is None
        rec = api.get(f"{BASE_URL}/api/records/{rid}").json()
        assert rec["folder_id"] in (None, "")

    def test_move_404_for_unknown(self, api, state):
        fid = state["folders"][0]
        r = api.post(f"{BASE_URL}/api/records/nope/move",
                     json={"folder_id": fid})
        assert r.status_code == 404


# ---------------------------------------------------------------------------
# 5. GET /api/records — q search across new fields
# ---------------------------------------------------------------------------
class TestSearchNewFields:
    def test_search_by_source(self, api, state):
        # record[0] source was updated to "Pune" during patch
        r = api.get(f"{BASE_URL}/api/records", params={"q": "Pune"})
        assert r.status_code == 200
        ids = [x["id"] for x in r.json()["items"]]
        assert state["records"][0] in ids

    def test_search_by_destination(self, api, state):
        r = api.get(f"{BASE_URL}/api/records", params={"q": "Delhi"})
        assert r.status_code == 200
        ids = [x["id"] for x in r.json()["items"]]
        assert state["records"][0] in ids

    def test_search_by_remarks(self, api, state):
        r = api.get(f"{BASE_URL}/api/records", params={"q": "edited-remark"})
        assert r.status_code == 200
        ids = [x["id"] for x in r.json()["items"]]
        assert state["records"][0] in ids

    def test_search_by_amount(self, api, state):
        r = api.get(f"{BASE_URL}/api/records", params={"q": "1600.00"})
        assert r.status_code == 200
        ids = [x["id"] for x in r.json()["items"]]
        assert state["records"][0] in ids

    def test_search_by_date(self, api, state):
        r = api.get(f"{BASE_URL}/api/records", params={"q": "2026-01-05"})
        assert r.status_code == 200
        ids = [x["id"] for x in r.json()["items"]]
        assert state["records"][0] in ids


# ---------------------------------------------------------------------------
# 6. GET /api/records — scope filters
# ---------------------------------------------------------------------------
class TestScopeFilters:
    def _rid_in(self, api, params, rid):
        r = api.get(f"{BASE_URL}/api/records", params=params)
        assert r.status_code == 200, r.text
        return rid in [x["id"] for x in r.json()["items"]]

    def test_scope_today(self, api, state):
        # Newly created records must show under scope=today
        assert self._rid_in(api, {"scope": "today"}, state["records"][0])

    def test_scope_week(self, api, state):
        assert self._rid_in(api, {"scope": "week"}, state["records"][0])

    def test_scope_month(self, api, state):
        assert self._rid_in(api, {"scope": "month"}, state["records"][0])

    def test_scope_year(self, api, state):
        assert self._rid_in(api, {"scope": "year"}, state["records"][0])

    def test_year_month(self, api, state):
        now = datetime.now(timezone.utc)
        assert self._rid_in(
            api, {"year": now.year, "month": now.month}, state["records"][0]
        )

    def test_year_only(self, api, state):
        now = datetime.now(timezone.utc)
        assert self._rid_in(api, {"year": now.year}, state["records"][0])

    def test_far_past_year_empty(self, api, state):
        # Records created now should NOT show in year=2000
        r = api.get(f"{BASE_URL}/api/records", params={"year": 2000})
        assert r.status_code == 200
        assert state["records"][0] not in [x["id"] for x in r.json()["items"]]


# ---------------------------------------------------------------------------
# 7. GET /api/folders/smart — smart counts + year/month breakdown
# ---------------------------------------------------------------------------
class TestSmartFolders:
    def test_smart_shape(self, api):
        r = api.get(f"{BASE_URL}/api/folders/smart")
        assert r.status_code == 200
        data = r.json()
        assert "smart" in data and "years" in data
        keys = {s["key"] for s in data["smart"]}
        assert {"today", "week", "month", "year", "all"}.issubset(keys)
        # All counts are non-negative ints
        for s in data["smart"]:
            assert isinstance(s["count"], int) and s["count"] >= 0
            assert "label" in s

    def test_smart_all_matches_db_total(self, api):
        smart = api.get(f"{BASE_URL}/api/folders/smart").json()
        stats = api.get(f"{BASE_URL}/api/stats").json()
        all_row = next(s for s in smart["smart"] if s["key"] == "all")
        assert all_row["count"] == stats["total_records"]

    def test_smart_today_gte_our_recent_creates(self, api, state):
        # We just inserted at least 2 records; today count must be >= 2
        smart = api.get(f"{BASE_URL}/api/folders/smart").json()
        today = next(s for s in smart["smart"] if s["key"] == "today")
        assert today["count"] >= 2

    def test_year_month_breakdown(self, api):
        now = datetime.now(timezone.utc)
        smart = api.get(f"{BASE_URL}/api/folders/smart").json()
        years = smart["years"]
        assert isinstance(years, list)
        this_year = next((y for y in years if y["year"] == f"{now.year:04d}"), None)
        assert this_year is not None, f"current year not in breakdown: {years}"
        assert this_year["total"] >= 2
        months = {m["month"] for m in this_year["months"]}
        assert f"{now.month:02d}" in months


# ---------------------------------------------------------------------------
# 8. POST /api/export — xlsx with scope filters, decode workbook
# ---------------------------------------------------------------------------
class TestExportXlsxScoped:
    def test_export_xlsx_all(self, api):
        r = api.post(f"{BASE_URL}/api/export", json={"format": "xlsx"})
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["filename"].endswith(".xlsx")
        assert d["mime_type"].startswith(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        raw = base64.b64decode(d["base64"])
        assert raw[:2] == b"PK", "not a valid xlsx zip file"

        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        header = [c.value for c in ws[1]]
        # Verify key fields present in header
        required_headers = {
            "id", "created_at", "verification_status",
            "document_number", "vehicle_number", "gstin",
            "source", "destination", "remarks", "amount",
        }
        missing = required_headers - set(header)
        assert not missing, f"missing xlsx headers: {missing}"

    def test_export_xlsx_scope_today_contains_our_record(self, api, state):
        r = api.post(f"{BASE_URL}/api/export",
                     json={"format": "xlsx", "scope": "today"})
        assert r.status_code == 200
        d = r.json()
        assert d["count"] >= 2  # our seeded records

        from openpyxl import load_workbook
        raw = base64.b64decode(d["base64"])
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        header = [c.value for c in ws[1]]
        id_col = header.index("id")
        doc_col = header.index("document_number")
        ids_in_sheet = []
        docs_in_sheet = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            ids_in_sheet.append(row[id_col])
            docs_in_sheet.append(row[doc_col])
        assert state["records"][0] in ids_in_sheet
        assert "TEST-STORE-1" in docs_in_sheet

    def test_export_xlsx_scope_year_month(self, api, state):
        now = datetime.now(timezone.utc)
        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "xlsx", "year": now.year, "month": now.month,
        })
        assert r.status_code == 200
        d = r.json()
        assert d["count"] >= 2
        raw = base64.b64decode(d["base64"])
        assert raw[:2] == b"PK"

    def test_export_xlsx_by_folder(self, api, state):
        # Move record back into the folder for this test
        rid = state["records"][0]
        fid = state["folders"][0]
        mv = api.post(f"{BASE_URL}/api/records/{rid}/move",
                      json={"folder_id": fid})
        assert mv.status_code == 200

        r = api.post(f"{BASE_URL}/api/export", json={
            "format": "xlsx", "folder_id": fid,
        })
        assert r.status_code == 200
        d = r.json()
        assert d["count"] >= 1

        from openpyxl import load_workbook
        raw = base64.b64decode(d["base64"])
        wb = load_workbook(io.BytesIO(raw))
        ws = wb.active
        header = [c.value for c in ws[1]]
        id_col = header.index("id")
        ids_in_sheet = [row[id_col] for row in ws.iter_rows(min_row=2, values_only=True)]
        assert rid in ids_in_sheet


# ---------------------------------------------------------------------------
# 9. Persistence — records survive multiple independent sessions
# ---------------------------------------------------------------------------
class TestPersistence:
    def test_records_persist_across_new_session(self, state):
        # Open a fresh HTTP session (no shared cookies/keepalive) to mimic
        # a restart-like independent client.
        rid = state["records"][0]
        s = requests.Session()
        s.headers.update({"Content-Type": "application/json"})
        r = s.get(f"{BASE_URL}/api/records/{rid}")
        assert r.status_code == 200
        assert r.json()["id"] == rid


# ---------------------------------------------------------------------------
# ZZ. Cleanup — delete records + folder created by this run
# ---------------------------------------------------------------------------
class TestZZCleanup:
    def test_delete_records(self, api, state):
        for rid in list(state["records"]):
            r = api.delete(f"{BASE_URL}/api/records/{rid}")
            assert r.status_code == 200
        state["records"].clear()

    def test_delete_folders(self, api, state):
        for fid in list(state["folders"]):
            r = api.delete(f"{BASE_URL}/api/folders/{fid}")
            assert r.status_code == 200
        state["folders"].clear()
